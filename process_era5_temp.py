#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Process ERA5 hourly 2m temperature data to calculate temperature changes
from a 1980-1989 baseline for each 5-year period.
Focus on 8 PM to midnight local time at each location.
"""

import os
import numpy as np
import pandas as pd
import xarray as xr
import geopandas as gpd
from datetime import datetime
import matplotlib.pyplot as plt
from pathlib import Path
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill

# Define paths
# Use absolute paths
BASE_DIR = Path(__file__).parent.absolute()
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# Define time periods
BASELINE_YEARS = list(range(1980, 1990))  # 1980-1989
PERIODS = [
    (1990, 1994),
    (1995, 1999),
    (2000, 2004),
    (2005, 2009),
    (2010, 2014),
    (2015, 2019),
    (2020, 2024),
]

# Define evening hours (8 PM to midnight)
EVENING_HOURS = list(range(20, 24))  # 20:00, 21:00, 22:00, 23:00

def load_era5_data(years):
    """
    Load ERA5 hourly 2m temperature data for the specified years.
    
    Args:
        years (list): List of years to load
        
    Returns:
        xarray.Dataset: Combined dataset for all years
    """
    print(f"Loading ERA5 data for years: {years}")
    
    # List of files to load
    files = [DATA_DIR / f"era5_hourly_2m_temperature_{year}.nc" for year in years]
    
    # Check if all files exist
    for file in files:
        if not file.exists():
            raise FileNotFoundError(f"File not found: {file}")
    
    # Load all files as a single dataset
    ds = xr.open_mfdataset(files, combine='by_coords')
    
    # Convert temperature from Kelvin to Celsius
    if 't2m' in ds:
        ds['t2m'] = ds['t2m'] - 273.15
    
    return ds

def extract_evening_hours(ds):
    """
    Extract data for evening hours (8 PM to midnight) in local time.
    
    Args:
        ds (xarray.Dataset): ERA5 dataset
        
    Returns:
        xarray.Dataset: Dataset with only evening hours
    """
    print("Extracting evening hours (8 PM to midnight) in local time using improved timezone mapping")
    
    # Create a copy of the dataset to avoid modifying the original
    ds_evening = ds.copy()
    
    # Get the hour of day in UTC
    hour_utc = ds_evening.valid_time.dt.hour
    
    # Define timezone boundaries for the continental US
    # These are approximate but more precise than longitude/15
    # Format: (min_longitude, max_longitude, offset)
    us_timezone_boundaries = [
        (-125.0, -115.0, -8),  # Pacific Time (UTC-8)
        (-115.0, -101.0, -7),  # Mountain Time (UTC-7)
        (-101.0, -87.0, -6),   # Central Time (UTC-6)
        (-87.0, -67.0, -5),    # Eastern Time (UTC-5)
    ]
    
    # Get longitude values
    longitudes = ds_evening.longitude.values
    
    # Create a mask for evening hours in local time
    masks = []
    
    # For each evening hour (8 PM to midnight)
    for hour in EVENING_HOURS:
        # Start with a mask of all False
        hour_mask = xr.zeros_like(hour_utc, dtype=bool)
        
        # For each timezone boundary
        for min_lon, max_lon, offset in us_timezone_boundaries:
            # Create a mask for longitudes in this timezone
            lon_mask = (longitudes >= min_lon) & (longitudes < max_lon)
            
            # If there are any points in this timezone
            if np.any(lon_mask):
                # Calculate the UTC hour that corresponds to the target local hour
                target_utc_hour = (hour - offset) % 24
                
                # Create a mask for times where UTC hour matches the target
                time_mask = abs((hour_utc - target_utc_hour) % 24) < 0.5
                
                # Update the hour mask for points in this timezone
                for i, is_in_tz in enumerate(lon_mask):
                    if is_in_tz:
                        lon_specific_mask = (ds_evening.longitude == longitudes[i])
                        hour_mask = hour_mask | (lon_specific_mask & time_mask)
        
        # For longitudes outside the defined boundaries, use longitude/15
        outside_tz_mask = ~np.logical_or.reduce([
            (longitudes >= min_lon) & (longitudes < max_lon) 
            for min_lon, max_lon, _ in us_timezone_boundaries
        ])
        
        if np.any(outside_tz_mask):
            for i, is_outside in enumerate(outside_tz_mask):
                if is_outside:
                    lon = longitudes[i]
                    # Calculate offset using longitude/15
                    offset = lon / 15
                    # Calculate the UTC hour that corresponds to the target local hour
                    target_utc_hour = (hour - offset) % 24
                    # Create a mask for this longitude
                    lon_specific_mask = (ds_evening.longitude == lon)
                    # Create a mask for times where UTC hour matches the target
                    time_mask = abs((hour_utc - target_utc_hour) % 24) < 0.5
                    # Update the hour mask
                    hour_mask = hour_mask | (lon_specific_mask & time_mask)
        
        masks.append(hour_mask)
    
    # Combine all masks with logical OR
    combined_mask = masks[0]
    for mask in masks[1:]:
        combined_mask = combined_mask | mask
    
    # Apply the mask to the dataset
    ds_evening = ds_evening.where(combined_mask, drop=True)
    
    return ds_evening

def calculate_hourly_averages(ds, local_hour=None):
    """
    Calculate average temperature for each hour of the day.
    
    Args:
        ds (xarray.Dataset): ERA5 dataset
        local_hour (int, optional): If provided, calculate average for this local hour only
        
    Returns:
        xarray.Dataset: Dataset with hourly averages
    """
    if local_hour is not None:
        print(f"Calculating average temperature for local hour {local_hour}")
    else:
        print("Calculating average temperature for each hour of the day")
    
    # Group by hour of day and calculate mean
    ds_hourly_avg = ds.groupby('valid_time.hour').mean()
    
    return ds_hourly_avg

def load_state_boundaries():
    """
    Load state boundaries from shapefile.
    
    Returns:
        geopandas.GeoDataFrame: State boundaries
    """
    print("Loading state boundaries")
    
    # Path to the shapefile
    shapefile_path = DATA_DIR / "ne_10m_admin_1_states_provinces" / "ne_10m_admin_1_states_provinces.shp"
    
    # Load the shapefile
    states = gpd.read_file(shapefile_path)
    
    # Filter to include only US states
    us_states = states[states['admin'] == 'United States of America']
    
    return us_states

def aggregate_by_state(ds, states):
    """
    Aggregate temperature data by state.
    
    Args:
        ds (xarray.Dataset): ERA5 dataset
        states (geopandas.GeoDataFrame): State boundaries
        
    Returns:
        pandas.DataFrame: Temperature data aggregated by state
    """
    print("Aggregating temperature data by state")
    
    # Convert dataset to a dataframe
    df = ds.to_dataframe().reset_index()
    
    # Create a GeoDataFrame with points
    gdf = gpd.GeoDataFrame(
        df, 
        geometry=gpd.points_from_xy(df.longitude, df.latitude),
        crs="EPSG:4326"  # Set the CRS to match the state boundaries
    )
    
    # Ensure both GeoDataFrames have the same CRS
    if gdf.crs != states.crs:
        print(f"Reprojecting points from {gdf.crs} to {states.crs}")
        gdf = gdf.to_crs(states.crs)
    
    # Spatial join with state boundaries
    joined = gpd.sjoin(gdf, states, how="inner", predicate="within")
    
    # Group by state and calculate mean
    state_avg = joined.groupby('name')['t2m'].mean().reset_index()
    
    return state_avg

def export_to_excel(state_data_dict, output_file):
    """
    Export state temperature change data to Excel.
    
    Args:
        state_data_dict (dict): Dictionary with period names as keys and DataFrames as values
        output_file (Path): Path to the output Excel file
    """
    print(f"Exporting data to Excel: {output_file}")
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # Create a summary sheet
    summary_sheet = wb.create_sheet("Summary")
    
    # Add header to summary sheet
    summary_sheet["A1"] = "State"
    summary_sheet["A1"].font = header_font
    summary_sheet["A1"].fill = header_fill
    
    # Add period headers to summary sheet
    for i, period in enumerate(state_data_dict.keys()):
        col = chr(66 + i)  # B, C, D, etc.
        summary_sheet[f"{col}1"] = period
        summary_sheet[f"{col}1"].font = header_font
        summary_sheet[f"{col}1"].fill = header_fill
        summary_sheet[f"{col}1"].alignment = center_alignment
    
    # Get all states from the first period
    first_period = list(state_data_dict.keys())[0]
    states = state_data_dict[first_period]["name"].tolist()
    
    # Add state data to summary sheet
    for row, state in enumerate(states, start=2):
        summary_sheet[f"A{row}"] = state
        
        for i, period in enumerate(state_data_dict.keys()):
            col = chr(66 + i)  # B, C, D, etc.
            df = state_data_dict[period]
            state_value = df[df["name"] == state]["t2m"].values[0]
            summary_sheet[f"{col}{row}"] = round(state_value, 2)
    
    # Create a sheet for each state
    for state in states:
        state_sheet = wb.create_sheet(state)
        
        # Add header to state sheet
        state_sheet["A1"] = "Period"
        state_sheet["B1"] = "Temperature Change (°C)"
        state_sheet["A1"].font = header_font
        state_sheet["B1"].font = header_font
        state_sheet["A1"].fill = header_fill
        state_sheet["B1"].fill = header_fill
        
        # Add period data to state sheet
        for row, period in enumerate(state_data_dict.keys(), start=2):
            df = state_data_dict[period]
            state_value = df[df["name"] == state]["t2m"].values[0]
            
            state_sheet[f"A{row}"] = period
            state_sheet[f"B{row}"] = round(state_value, 2)
    
    # Save the workbook
    wb.save(output_file)
    print(f"Excel file saved: {output_file}")

def main():
    """Main function to process ERA5 data."""
    print("Starting ERA5 data processing")
    
    # Create output directory if it doesn't exist
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    # Load state boundaries
    states = load_state_boundaries()
    
    # Process baseline period (1980-1989)
    print("\nProcessing baseline period (1980-1989)")
    baseline_ds = load_era5_data(BASELINE_YEARS)
    baseline_evening_ds = extract_evening_hours(baseline_ds)
    baseline_hourly_avg = calculate_hourly_averages(baseline_evening_ds)
    
    # Save baseline hourly averages
    baseline_hourly_avg.to_netcdf(OUTPUT_DIR / "baseline_hourly_avg.nc")
    print(f"Saved baseline hourly averages to {OUTPUT_DIR / 'baseline_hourly_avg.nc'}")
    
    # Dictionary to store state temperature change data for each period
    state_temp_changes = {}
    
    # Process each 5-year period
    for start_year, end_year in PERIODS:
        period_years = list(range(start_year, end_year + 1))
        period_name = f"{start_year}-{end_year}"
        
        print(f"\nProcessing period: {period_name}")
        
        # Load data for this period
        period_ds = load_era5_data(period_years)
        period_evening_ds = extract_evening_hours(period_ds)
        period_hourly_avg = calculate_hourly_averages(period_evening_ds)
        
        # Save period hourly averages
        period_hourly_avg.to_netcdf(OUTPUT_DIR / f"{period_name}_hourly_avg.nc")
        print(f"Saved period hourly averages to {OUTPUT_DIR / f'{period_name}_hourly_avg.nc'}")
        
        # Calculate temperature change from baseline
        temp_change = period_hourly_avg - baseline_hourly_avg
        
        # Save temperature change
        temp_change.to_netcdf(OUTPUT_DIR / f"{period_name}_temp_change.nc")
        print(f"Saved temperature change to {OUTPUT_DIR / f'{period_name}_temp_change.nc'}")
        
        # Aggregate temperature change by state
        state_avg = aggregate_by_state(temp_change, states)
        
        # Store state temperature change data
        state_temp_changes[period_name] = state_avg
    
    # Export state temperature change data to Excel
    excel_output = OUTPUT_DIR / "temperature_changes_by_state.xlsx"
    export_to_excel(state_temp_changes, excel_output)
    
    print("\nProcessing complete!")

if __name__ == "__main__":
    main()
